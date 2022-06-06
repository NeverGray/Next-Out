from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

import NV_CONSTANTS
import NV_run

SHEET_NAMES ={
    "SSA" :"Second-by-second Aerodynamic Data (SSA)",
    "SST" :"Second-by-second Thermodynamic data (SST)",
    "TRA" :"Second-by-second Train Data (TRA)",
    "SA"  :"Summary of Aerodynamic Data (SA)",
    "ST"  :"Summary of Thermodynamic Data (ST)",
    "PER" :"Percentage of Time Temperature is Above (PER)",
    "TES" :"Train Energy Summary (TES)",
    "HSA" :"Heat Sink Analysis (for uncontrolled zones) (HSA)",
    "ECS" :"Environmental Control System Load Estimates (ECS)",
    "SA-" :"Summary of Aerodynamic Data (SA)",
    "ST-" :"Summary of Thermodynamic Data (ST)"
}

def create_excel(settings, data, output_meta_data, gui=""):
    file_name = str(output_meta_data['file_path'].name)
    excel_results_path = NV_run.get_results_path2(settings, output_meta_data, ".xlsx")
    # TODO Add error checker if excel file is open
    NV_run.run_msg(gui, "Creating Excel file " + excel_results_path.name)
    TITLES = {
        "File Name:" : output_meta_data['file_path'].name,
        "File Time:" : output_meta_data['file_time'],
        "Data:": "From worksheet name",
        "Units:": output_meta_data['ses_version']
        }
    #Select index for units in NV_CONSTANTS.COLUMN_UNITS value
    if output_meta_data['ses_version'] == 'IP':
        unit_index = 1
    else:
        unit_index = 0
    try:
        bio = BytesIO()
        with pd.ExcelWriter(bio, engine="openpyxl") as writer:
            for item in data.values():
                item.to_excel(writer, sheet_name=item.name, merge_cells=False)
                worksheet = writer.sheets[item.name]
                # Create title rows and format
                df_startrow = len(TITLES) + 1
                worksheet.insert_rows(0,df_startrow - 1)
                i = 1
                for x, y in TITLES.items():
                    worksheet.cell(row=i,column = 1, value=x).font = Font(size=10, bold=True)
                    worksheet.cell(row=i,column = 2, value=y)
                    i +=1
                worksheet.cell(row=df_startrow-2, column =2, value = SHEET_NAMES.get(item.name[:3]))
                # Create filters from https://stackoverflow.com/questions/51566349/openpyxl-how-to-add-filters-to-all-columns
                df_dimensions = 'A'+ str(df_startrow) + worksheet.dimensions[2:]
                worksheet.auto_filter.ref = df_dimensions
                # Freeze cells from https://stackoverflow.com/questions/25588918/how-to-freeze-entire-header-row-in-openpyxl
                freeze_column = len(item.index.names)
                freeze_cell = worksheet.cell(row=df_startrow + 1, column=freeze_column + 1)
                worksheet.freeze_panes = freeze_cell
                #Rotate Header rows
                for col in worksheet.iter_cols(min_row=df_startrow, max_row=df_startrow, min_col=len(item.index.names)+1):
                    for cell in col:
                        cell.alignment = Alignment(textRotation= 45)
                #Format top row
                maximum_column_number =  len(item.index.names) + len(item.columns)
                if output_meta_data['ses_version'] == "IP_TO_SI":
                    color_code = "4BACC6" #Blue
                else:
                    color_code = "F79646" #Orange
                for row in worksheet.iter_rows(min_row=df_startrow, max_col=maximum_column_number, max_row=df_startrow):
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor=color_code)
                #Add units to top row
                unit_row = 4
                name_row = 5
                unit_column_start = freeze_column + 1
                unit_column_end = unit_column_start + len(item.columns)
                for col in range(unit_column_start, unit_column_end):
                    column_name = worksheet.cell(row=name_row,column=col).value
                    cell = worksheet.cell(row=unit_row, column=col)
                    cell.alignment = Alignment(horizontal='center')
                    if column_name in NV_CONSTANTS.COLUMN_UNITS:
                        cell.value = NV_CONSTANTS.COLUMN_UNITS[column_name][unit_index]
            # Add properties to Excel File.  Following https://stackoverflow.com/questions/52120125/how-to-edit-core-properties-of-xlsx-file-with-python'''
            writer.book.properties.creator = ("Next Vis" + NV_CONSTANTS.VERSION_NUMBER)
            writer.book.properties.title = file_name
            writer.save()
            # From https://techoverflow.net/2019/07/24/how-to-write-bytesio-content-to-file-in-python/

            with open(excel_results_path, "wb") as outfile:  
                    # Copy the BytesIO stream to the output file
                    try:
                        outfile.write(bio.getvalue())
                    except:
                        print(
                            "Error writing "
                            + file_name
                            + ".xlsx. Try closing file and trying again."
                        )
        NV_run.run_msg(gui, "Created Excel file " + excel_results_path.name)
    except:
        NV_run.run_msg(gui, "ERROR creating Excel file "+ excel_results_path.name + ". Try closing file and process again.")

if __name__ == "__main__":
    file_path_string = "C:/Simulations/Next-Vis/Excel Speedup/sinorm-detailed.OUT"
    visio_template = "C:/Users/msn/OneDrive - Never Gray/Software Development/Next-Vis/Python2021/sample012.vsdx"
    results_folder_str = "C:/Simulations/Next-Vis/Excel Speedup"
    settings = {
        "ses_output_str": [file_path_string],
        "visio_template": visio_template,
        "results_folder_str": results_folder_str,
        "simtime": 9999.0,
        "version": "IP_TO_SI",
        "control": "First",
        "output": ["Excel"],
    }
    import cProfile
    import time

    import NV_parser
    file_path = Path(settings['ses_output_str'][0])
    data, output_meta_data = NV_parser.parse_file(file_path)
    file_name = file_path.name
    start_create_excel = time.perf_counter()
    create_excel(settings, data, output_meta_data)
    # cProfile.run('create_excel(settings, data, output_meta_data)')
    end_create_excel = time.perf_counter()
    print(f"Time for create_excel {end_create_excel - start_create_excel:0.4f} seconds")
    # NV_run.single_sim(settings)
