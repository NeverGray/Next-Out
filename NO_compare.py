# Project Name: Next-Out
# Description: Compare two SES output files.
# Copyright (c) 2024 Justin Edenbaum, Never Gray
#
# This file is licensed under the MIT License.
# You may obtain a copy of the license at https://opensource.org/licenses/MIT

from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font

import NO_Excel_R01 as NV_excel
import NO_parser
import NO_run


def compare_outputs(settings, gui=""):
    if "Excel" in settings['output']: 
        Excel = True
    else:
        Excel = False
    base_file_str = settings['ses_output_str'][0]
    second_file_str = settings['ses_output_str'][1]
    #IF these are input files, perform SES simulations
    if settings["file_type"] == "input_file":
        base_file_str = NO_run.average_or_compare_call_ses(settings, base_file_str, gui)
        if base_file_str == 'Simulation failed':
            msg = 'Simulation failed'
            NO_run.run_msg(gui, msg)
            return
        second_file_str = NO_run.average_or_compare_call_ses(settings, second_file_str, gui)
        if base_file_str == 'Simulation failed':
            msg = 'Simulation failed'
            NO_run.run_msg(gui, msg)
            return
    base_file = Path(base_file_str)
    second_file = Path(second_file_str)
    base_df, base_output_meta_data = NO_parser.parse_file(base_file, gui, settings['conversion'])
    second_df, second_output_meta_data = NO_parser.parse_file(second_file, gui, settings['conversion'])
    if Excel:
        NV_excel.create_excel(settings, base_df, base_output_meta_data, gui)
        NV_excel.create_excel(settings, second_df, second_output_meta_data, gui)   
    base_data = dictionary_to_list(base_df)
    second_data = dictionary_to_list(second_df)
    num_df = len(base_data)
    suffix = base_output_meta_data['file_path'].suffix
    base_path = NO_run.get_results_path2(settings, base_output_meta_data, suffix)
    suffix = second_output_meta_data['file_path'].suffix
    second_path = NO_run.get_results_path2(settings, second_output_meta_data, suffix)
    if num_df != len(second_data):
        msg = "Error in Comparing two output files! " + base_path.name + "and" + second_path.name + "have different structures."
        NO_run.run_msg(gui, msg)
    else:
        msg = f'Comparing {base_path.name} and {second_path.name}.'
        NO_run.run_msg(gui, msg)
        diff = []
        diff_summary = []
        p_e = []  # percent Error
        p_e_summary = []
        base_data = remove_columns(base_data)
        second_data = remove_columns(second_data)
        for i in range(num_df):
            diff_data = second_data[i] - base_data[i]
            diff_data.name = base_data[i].name
            diff.append(diff_data)
            sum_data = diff_data.describe()
            sum_data.name = diff_data.name
            diff_summary.append(sum_data)
            p_e_data = abs(diff_data / base_data[i])
            p_e_data.name = base_data[i].name
            p_e.append(p_e_data)
            p_e_sum_data = p_e_data.describe()
            p_e_sum_data.name = diff_data.name
            p_e_summary.append(p_e_sum_data)
        try:
            # Description of Stem from https://automatetheboringstuff.com/2e/chapter9/
            file_name = base_path.stem + "_to_" + second_path.stem
            file_name = file_name[0:250] #make name isn't too long
            file_name_4_path = file_name + ".xlsx"
            ses_output_str = settings['ses_output_str'][0]
            parent = str(Path(ses_output_str).parent)
            base_output_meta_data['file_path'] = Path(parent + '/' + file_name_4_path)
            output_meta_data = base_output_meta_data.copy()
            output_meta_data['ses_version'] = 'Unconfirmed'
            compare_results_path = NO_run.get_results_path2(settings, output_meta_data, ".xlsx")
            p_e_text = (
                "Percent Error = Absolute value of [(Difference) / ("
                + base_path.name
                + ")]"
            )
            diff_text = "Difference = (" + second_path.name + ") - (" + base_path.name + ")"
            bio = BytesIO()
            # with pd.ExcelWriter(file_name + ".xlsx") as writer:
            with pd.ExcelWriter(bio, engine="openpyxl") as writer:
                # Code based on https://stackoverflow.com/questions/32957441/putting-many-python-pandas-dataframes-to-one-excel-worksheet
                for i in range(len(p_e_summary)):
                    sum_col = len(diff[i].index.names) - 1  # Aligns summary column
                    data_col = len(diff[i].index.names) + len(diff[i].columns)
                    s_name = p_e_summary[i].name
                    title = "Next-Out Sheet: " + s_name
                    n = 0
                    start_summary_row = 3
                    data_row = start_summary_row + 11  # Starts row
                    p_e_summary[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=start_summary_row,
                        startcol=sum_col,
                    )
                    p_e[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=0,
                    )
                    ws = writer.sheets[s_name]
                    ws.cell(row=1, column=1, value=title)
                    ws.cell(row=1, column=1).font = Font(bold=True)
                    ws.cell(row=2, column=1, value="Summary")
                    ws.cell(row=2, column=1).font = Font(underline="single")
                    ws.cell(row = data_row -1, column=1, value="Data")
                    ws.cell(row = data_row -1, column=1).font = Font(underline="single")
                    ws.cell(row=3, column=sum_col, value=p_e_text)
                    ws.cell(row=data_row, column=1, value=p_e_text)
                    n += 1
                    startcol_num = data_col * n + n
                    diff_summary[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=start_summary_row,
                        startcol=sum_col + startcol_num,
                    )
                    diff[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=startcol_num,
                    )
                    n = 1
                    startcol_num = (data_col * n + n) + 1
                    ws.cell(row=3, column=startcol_num, value=diff_text)
                    ws.cell(row=data_row, column=startcol_num, value=diff_text)
                    n += 1
                    startcol_num = data_col * n + n
                    second_data[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=startcol_num,
                    )
                    startcol_num = (data_col * n + n) + 1
                    ws.cell(row=data_row, column=startcol_num, value=second_path.name)
                    n += 1
                    startcol_num = data_col * n + n
                    base_data[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=startcol_num,
                    )
                    startcol_num = (data_col * n + n) + 1
                    ws.cell(row=data_row, column=startcol_num, value=base_path.name)
                    #Freeze Plane
                    freeze_cell = ws.cell(row=data_row+2, column=len(p_e[i].index.names)+1)
                    ws.freeze_panes = freeze_cell
                writer.book.properties.creator = "Next Vis 1p11`"
                writer.book.properties.title = file_name
                writer.book.save(bio) #Updated after Next-Vis 1p31
                # From https://techoverflow.net/2019/07/24/how-to-write-bytesio-content-to-file-in-python/
                # Copy the BytesIO stream to the output file
                with open(compare_results_path, "wb") as outfile:  
                    try:
                        outfile.write(bio.getvalue())
                    except:
                        NO_run.run_msg(gui,
                            "Error writing "
                            + str(compare_results_path)
                            + ".xlsx. Try closing file and trying again."
                        )
                    # TODO Add strings using https://stackoverflow.com/questions/43537598/write-strings-text-and-pandas-dataframe-to-excel
            msg = ("Created " + str(compare_results_path))
            NO_run.run_msg(gui, msg)
        except:
            msg = (
                "CRITICAL ERROR! Constructing (not saving) Excel File "
                + file_name
                + ".xlsx. Close the file if opened."
            )
            NO_run.run_msg(gui, msg)

def dictionary_to_list(dic):
    new_list = []
    for value in dic.values():
        new_list.append(value)
    return new_list

def remove_columns(df_list):
    #Removes columns that are strings so comparision can be completed.
    df_list[0].drop(['ID','Title'], axis =1, inplace = True)
    df_list[1].drop('ID', axis =1, inplace = True)
    return df_list

if __name__ == "__main__":
    directory_str = 'C:/Simulations/OpenSES_PR22_Tests/'
    ses_output_list = [
        directory_str + 'inferno4p2.out', 
        directory_str + 'inferno4p2_pr22.out'
        ]
    settings = {
        "ses_output_str": ses_output_list,
        "results_folder_str": directory_str,
        "visio_template": None,
        "simtime": 9999.0,
        "conversion": "",
        "control": "First",
        "output": ["Excel"],
        "file_type": "output_file",
        "path_exe": "C:\\simulations\\_EXE\\SES41.exe"
    }
    import datetime
    now1 = datetime.datetime.now()
    print ("start date and time : ")
    print (now1.strftime("%Y-%m-%d %H:%M:%S"))
    compare_outputs(settings)
    now2 = datetime.datetime.now()
    print ("end date and time : ")
    print (now2.strftime("%Y-%m-%d %H:%M:%S"))
    print(now2-now1)

