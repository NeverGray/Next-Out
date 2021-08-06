from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font

import NV_file_manager as nfm
import NV_parser as nvp


def compare_outputs(base_file="", second_file="",settings={}):
    base_list = nvp.parse_file(base_file)
    second_list = nvp.parse_file(second_file)
    base_data = base_list[0]
    second_data = second_list[0]
    print("Both Files post-processed")
    num_df = len(base_data)
    if num_df != len(second_data):
        print("Error - Parsed files are different")
    else:
        diff = []
        diff_summary = []
        p_e = []  # percent Error
        p_e_summary = []
        base_data = remove_columns(base_data)
        second_data = remove_columns(second_data)
        for i in range(num_df):
            diff_data = second_data[i] - base_data[i]
            diff_data.name = base_data[i].name
            diff.append(diff_data)
            sum_data = diff_data.describe()
            sum_data.name = diff_data.name
            diff_summary.append(sum_data)
            p_e_data = abs(diff_data / base_data[i]) * 100
            p_e_data.name = base_data[i].name
            p_e.append(p_e_data)
            p_e_sum_data = p_e_data.describe()
            p_e_sum_data.name = diff_data.name
            p_e_summary.append(p_e_sum_data)
        try:
            # Description of Stem from https://automatetheboringstuff.com/2e/chapter9/
            base_name = base_file.name 
            second_name = second_file.name
            file_name = base_file.stem + "_to_" + second_file.stem
            file_name = file_name[0:250] #make name isn't too long
            file_name_4_path = file_name + ".xlsx"
            #TODO update to use 'settings' and select output folder
            compare_results_path = second_file.parent/Path(file_name_4_path)
            p_e_text = (
                "Percent Error = Absolute value of [(Difference) / ("
                + base_name
                + ")] x 100%"
            )
            diff_text = "Difference = (" + second_name + ") - (" + base_name + ")"
            bio = BytesIO()
            # with pd.ExcelWriter(file_name + ".xlsx") as writer:
            with pd.ExcelWriter(bio, engine="openpyxl") as writer:
                # Code based on https://stackoverflow.com/questions/32957441/putting-many-python-pandas-dataframes-to-one-excel-worksheet
                for i in range(len(p_e_summary)):
                    sum_col = len(diff[i].index.names) - 1  # Aligns summary column
                    data_col = len(diff[i].index.names) + len(diff[i].columns)
                    s_name = p_e_summary[i].name
                    title = "Next-Vis Sheet: " + s_name
                    n = 0
                    start_summary_row = 3
                    data_row = start_summary_row + 11  # Starts row
                    p_e_summary[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=start_summary_row,
                        startcol=sum_col,
                    )
                    p_e[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=0,
                    )
                    ws = writer.sheets[s_name]
                    ws.cell(row=1, column=1, value=title)
                    ws.cell(row=1, column=1).font = Font(bold=True)
                    ws.cell(row=2, column=1, value="Summary")
                    ws.cell(row=2, column=1).font = Font(underline="single")
                    ws.cell(row = data_row -1, column=1, value="Data")
                    ws.cell(row = data_row -1, column=1).font = Font(underline="single")
                    ws.cell(row=3, column=sum_col, value=p_e_text)
                    ws.cell(row=data_row, column=1, value=p_e_text)
                    n += 1
                    startcol_num = data_col * n + n
                    diff_summary[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=start_summary_row,
                        startcol=sum_col + startcol_num,
                    )
                    diff[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=startcol_num,
                    )
                    n = 1
                    startcol_num = (data_col * n + n) + 1
                    ws.cell(row=3, column=startcol_num, value=diff_text)
                    ws.cell(row=data_row, column=startcol_num, value=diff_text)
                    n += 1
                    startcol_num = data_col * n + n
                    second_data[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=startcol_num,
                    )
                    startcol_num = (data_col * n + n) + 1
                    ws.cell(row=data_row, column=startcol_num, value=second_name)
                    n += 1
                    startcol_num = data_col * n + n
                    base_data[i].to_excel(
                        writer,
                        sheet_name=s_name,
                        merge_cells=False,
                        startrow=data_row,
                        startcol=startcol_num,
                    )
                    startcol_num = (data_col * n + n) + 1
                    ws.cell(row=data_row, column=startcol_num, value=base_name)
                    writer.book.properties.creator = "Next Vis Beta"
                    writer.book.properties.title = file_name
                writer.save()
                # From https://techoverflow.net/2019/07/24/how-to-write-bytesio-content-to-file-in-python/
                # Copy the BytesIO stream to the output file
                with open(compare_results_path, "wb") as outfile:  
                    try:
                        outfile.write(bio.getvalue())
                    except:
                        print(
                            "Error writing "
                            + str(compare_results_path)
                            + ".xlsx. Try closing file and trying again."
                        )
                    # TODO Add strings using https://stackoverflow.com/questions/43537598/write-strings-text-and-pandas-dataframe-to-excel
            print("Created Excel File " + str(compare_results_path))
            # TODO Remove or enable "Repeat option"
        except:
            print(
                "CRITICAL ERROR! Constructing (not saving) Excel File"
                + file_name
                + ".xlsx."
            )

def remove_columns(df_list):
    #Removes columns that are strings so comparision can be completed.
    df_list[0].drop(['ID','Title'], axis =1, inplace = True)
    df_list[1].drop('ID', axis =1, inplace = True)
    return df_list

if __name__ == "__main__":
    base_path_str = ['C:\\Temp\\4p2\\inferno.PRN','C:\\Temp\\4p2\\normal.PRN']
    second_path_str = ['C:\\Temp\\4p2\\inferno4p2.out','C:\\Temp\\4p2\\normal4p2.out']
    for i in range(len(base_path_str)):
        #if i ==1: break #for testing one

        compare_outputs(base_file=Path(base_path_str[i]), second_file=Path(second_path_str[i]))
